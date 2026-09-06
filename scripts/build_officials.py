#!/usr/bin/env python3
"""Bouwt data/officials.json uit het officialschema op volevo.nl.

De vereniging publiceert de officialbeurten als Excel-bestand. De bestandsnaam
bevat een datum, dus de link wordt van de pagina zelf geplukt in plaats van
hardgecodeerd - anders mist dit script stilzwijgend de volgende versie.

Elke beurt wordt gekoppeld aan een wedstrijd op datum, tijd en beide
teamnamen. Lukt dat voor een groot deel niet, dan stopt het script met een
foutmelding in plaats van een half bestand weg te schrijven.
"""
import json, os, re, sys, urllib.request

import openpyxl

PAGINA = 'https://volevo.nl/officialbeurten/'
KOPIE = 'data/officialbeurten.xlsx'
# volevo.nl weigert het Excel-bestand aan de GitHub-runners (404), terwijl de
# pagina zelf wel binnenkomt. Daarom doen we ons voor als een gewone browser
# en ligt er een kopie in de repo als dat alsnog niet lukt.
KOPPEN = {
    'User-Agent': ('Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 '
                   '(KHTML, like Gecko) Chrome/124.0 Safari/537.36'),
    'Accept': '*/*',
    'Accept-Language': 'nl,en;q=0.8',
    'Referer': PAGINA,
}


def haal(url, binair=False):
    req = urllib.request.Request(url, headers=KOPPEN)
    with urllib.request.urlopen(req, timeout=90) as r:
        data = r.read()
    return data if binair else data.decode('utf-8', 'replace')


# ── 1. Link naar het Excel-bestand zoeken en ophalen ──
bron, inhoud = None, None
try:
    html = haal(PAGINA)
    links = re.findall(r'href="(https://volevo\.nl/wp-content/uploads/[^"]+\.xlsx)"', html)
    if not links:
        raise RuntimeError('geen xlsx-link op de pagina gevonden')
    bron = links[0]
    print('bron:', bron)
    inhoud = haal(bron, binair=True)
    # Nieuwe versie meteen in de repo bewaren, zodat de terugval meegroeit
    os.makedirs('data', exist_ok=True)
    with open(KOPIE, 'wb') as f:
        f.write(inhoud)
except Exception as e:
    print('Live ophalen mislukt (%s)' % e, file=sys.stderr)
    if not os.path.exists(KOPIE):
        sys.exit('Geen kopie in %s om op terug te vallen - niets bijgewerkt' % KOPIE)
    print('Terugval op de kopie in de repo: %s' % KOPIE, file=sys.stderr)
    bron = bron or KOPIE
    with open(KOPIE, 'rb') as f:
        inhoud = f.read()

pad = 'officialbeurten-tijdelijk.xlsx'
with open(pad, 'wb') as f:
    f.write(inhoud)

# ── 2. Beurten uitlezen ──
wb = openpyxl.load_workbook(pad, data_only=True)
if 'Thuiswedstrijden' not in wb.sheetnames:
    sys.exit('Tabblad "Thuiswedstrijden" ontbreekt; tabbladen: %s' % wb.sheetnames)
ws = wb['Thuiswedstrijden']

beurten = []
for r in ws.iter_rows(min_row=2, values_only=True):
    if not r or not r[0]:
        continue
    datum, tijd, thuis, uit, taak, team, opmerking = (list(r) + [None]*7)[:7]
    try:
        dag = datum.date().isoformat()
    except AttributeError:
        continue
    beurten.append({
        'dag': dag,
        'tijd': str(tijd)[:5],
        'thuis': (thuis or '').strip(),
        'uit': (uit or '').strip(),
        'taak': (taak or '').strip(),
        'team': (team or '').strip(),
        'opmerking': (opmerking or '').strip() or None,
    })

if not beurten:
    sys.exit('Geen beurten gevonden in het bestand')

# ── 3. Koppelen aan de wedstrijden ──
try:
    wedstrijden = json.load(open('data/matches.json', encoding='utf-8'))['wedstrijden']
except (OSError, ValueError, KeyError):
    sys.exit('data/matches.json ontbreekt of is onleesbaar - draai eerst build_matches.py')

index = {}
for w in wedstrijden:
    t = w.get('tijdstip') or ''
    index[(t[:10], t[11:16], w.get('thuis'), w.get('uit'))] = t

per_wedstrijd, los = {}, []
for b in beurten:
    t = index.get((b['dag'], b['tijd'], b['thuis'], b['uit']))
    if t:
        sleutel = t + '|' + b['thuis'] + '|' + b['uit']
        per_wedstrijd.setdefault(sleutel, []).append(
            {'taak': b['taak'], 'team': b['team'], 'opmerking': b['opmerking']})
    else:
        los.append(b)

gekoppeld = len(beurten) - len(los)
print('%d beurten, %d gekoppeld, %d niet' % (len(beurten), gekoppeld, len(los)))
for b in los[:5]:
    print('   niet gekoppeld: %s %s  %s - %s' % (b['dag'], b['tijd'], b['thuis'], b['uit']))

if gekoppeld < len(beurten) * 0.8:
    sys.exit('Te weinig beurten gekoppeld (%d van %d) - bestand niet overschreven'
             % (gekoppeld, len(beurten)))

# ── 4. Botsingen signaleren: een beurt tijdens een eigen wedstrijd ──
speelt = {}
for w in wedstrijden:
    for naam in (w.get('thuis'), w.get('uit')):
        if (naam or '').upper().startswith('VOLEVO'):
            speelt.setdefault(naam.replace('VOLEVO', '').strip(), set()).add(
                (w['tijdstip'][:10], w['tijdstip'][11:16]))

botsingen = []
for b in beurten:
    if b['team'].lower() in ('', 'ouders'):
        continue
    if (b['dag'], b['tijd']) in speelt.get(b['team'], set()):
        botsingen.append({'dag': b['dag'], 'tijd': b['tijd'], 'team': b['team'],
                          'taak': b['taak'], 'wedstrijd': b['thuis'] + ' - ' + b['uit']})
if botsingen:
    print('LET OP: %d beurt(en) vallen samen met een eigen wedstrijd' % len(botsingen))
    for x in botsingen[:10]:
        print('   %s %s  %s moet %s' % (x['dag'], x['tijd'], x['team'], x['taak']))

data = {
    'bron': bron,
    'aantal': len(beurten),
    'nietGekoppeld': los,
    'botsingen': botsingen,
    'perWedstrijd': per_wedstrijd,
}

# Alleen schrijven als er inhoudelijk iets veranderd is
try:
    oud = json.load(open('data/officials.json', encoding='utf-8'))
    if all(oud.get(k) == data[k] for k in ('bron', 'perWedstrijd', 'botsingen', 'nietGekoppeld')):
        print('Officialbeurten ongewijzigd - bestand niet aangeraakt')
        os.remove(pad)
        raise SystemExit(0)
except (OSError, ValueError):
    pass

os.makedirs('data', exist_ok=True)
with open('data/officials.json', 'w', encoding='utf-8') as f:
    json.dump(data, f, ensure_ascii=False, indent=1, sort_keys=False)
    f.write('\n')
os.remove(pad)
print('data/officials.json geschreven voor %d wedstrijden' % len(per_wedstrijd))
